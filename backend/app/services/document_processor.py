"""
Document processing service for FinMDA-Bot.
"""
from typing import Optional
from datetime import datetime
import io

from app.models import Document
from app.database import SessionLocal
from app.services.rag_service import RAGService

try:
    import fitz  # PyMuPDF
except Exception:
    fitz = None

try:
    import openpyxl
except Exception:
    openpyxl = None


class DocumentProcessor:
    """Service class for processing uploaded documents."""

    async def process_document(self, document_id: int, file_path: str, file_type: str) -> None:
        """Extract lightweight text and index into RAG (best-effort)."""
        extracted_text = ""
        metadata = {
            "processing_timestamp": datetime.utcnow().isoformat(),
            "file_type": file_type,
        }

        try:
            ft = (file_type or "").lower()
            if ft == "pdf" and fitz is not None:
                doc = fitz.open(file_path)
                texts = []
                for page in doc:
                    texts.append(page.get_text())
                extracted_text = "\n\n".join(texts)
                metadata["page_count"] = len(doc)
                doc.close()
            elif ft in ("csv",):
                # Read as text; limit size to avoid memory blowups
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    extracted_text = f.read(200_000)
            elif ft in ("xlsx", "xls") and openpyxl is not None:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                metadata["sheets"] = wb.sheetnames
                # Grab first sheet preview
                sheet = wb[wb.sheetnames[0]]
                rows = []
                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    if i >= 50:
                        break
                    cleaned = ["" if c is None else str(c) for c in row]
                    rows.append(",".join(cleaned))
                extracted_text = "\n".join(rows)
                wb.close()
            else:
                # Fallback: no extraction
                extracted_text = ""
        except Exception as e:
            # Best-effort extraction; continue
            metadata["extraction_error"] = str(e)

        # Persist results and best-effort index into RAG
        with SessionLocal() as db:
            document = db.query(Document).filter(Document.id == document_id).first()
            if not document:
                return
            document.extracted_text = extracted_text[:1_000_000] if extracted_text else None
            document.document_metadata = metadata
            document.is_processed = True
            db.commit()

        # Index into Chroma (ignore failures)
        if extracted_text:
            try:
                rag = RAGService()
                await rag.index_document(document_id, extracted_text, metadata)
            except Exception:
                pass
